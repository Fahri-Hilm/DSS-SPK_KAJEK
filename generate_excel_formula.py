import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load data
df = pd.read_csv('No-Vendor-NamaPaketPlan-CPU-RAM-DiskIOSpeed-HargaBulanUSD.csv')

# Extract values
df['CPU_val'] = df['CPU'].apply(lambda x: int(re.search(r'(\d+)', x).group(1)))
df['RAM_val'] = df['RAM'].apply(lambda x: int(re.search(r'(\d+)', x).group(1)))

def extract_disk_io(val):
    speed_match = re.search(r'(\d+(?:\.\d+)?)\s*(Gbps|Mbps)', val)
    if speed_match:
        speed = float(speed_match.group(1))
        if 'Gbps' in speed_match.group(2):
            speed *= 1000
    else:
        speed = 100.0
    return speed / 8

df['DiskIO_val'] = df['Disk I/O Speed'].apply(extract_disk_io)
df['Price_val'] = df['Harga/Bulan (USD)'].apply(lambda x: float(re.search(r'[\$\s]*([\d.]+)', x).group(1)))

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E78")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Sheet 1: Data Input
ws1 = wb.create_sheet("1. Data Input")
ws1['A1'] = "DATA INPUT - UBAH NILAI DI SINI"
ws1['A1'].font = title_font

headers = ['No', 'Vendor', 'Nama Paket', 'CPU', 'RAM', 'Disk I/O', 'Harga']
for idx, h in enumerate(headers, 1):
    cell = ws1.cell(3, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Input data with yellow background for editable cells
for idx, row in df.iterrows():
    ws1.cell(idx+4, 1, row['No'])
    ws1.cell(idx+4, 2, row['Vendor'])
    ws1.cell(idx+4, 3, row['Nama Paket (Plan)'])
    
    # Editable cells (yellow)
    for col in [4, 5, 6, 7]:
        cell = ws1.cell(idx+4, col)
        cell.fill = yellow_fill
    
    ws1.cell(idx+4, 4, row['CPU_val'])
    ws1.cell(idx+4, 5, row['RAM_val'])
    ws1.cell(idx+4, 6, round(row['DiskIO_val'], 2))
    ws1.cell(idx+4, 7, round(row['Price_val'], 2))

ws1['A25'] = "⚠️ UBAH NILAI DI KOLOM KUNING (D-G) UNTUK RECALCULATE OTOMATIS"
ws1['A25'].font = Font(bold=True, color="FF0000", size=12)

# Sheet 2: Bobot & Kriteria
ws2 = wb.create_sheet("2. Bobot & Kriteria")
ws2['A1'] = "BOBOT DAN TIPE KRITERIA"
ws2['A1'].font = title_font

ws2['A3'] = "Kriteria"
ws2['B3'] = "CPU (C1)"
ws2['C3'] = "RAM (C2)"
ws2['D3'] = "Disk I/O (C3)"
ws2['E3'] = "Harga (C4)"
for cell in ws2[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws2['A4'] = "Tipe"
ws2['B4'] = "BENEFIT"
ws2['C4'] = "BENEFIT"
ws2['D4'] = "BENEFIT"
ws2['E4'] = "COST"

ws2['A5'] = "Bobot (W)"
for col in ['B5', 'C5', 'D5', 'E5']:
    ws2[col] = 0.25
    ws2[col].fill = yellow_fill

ws2['A7'] = "Total Bobot:"
ws2['B7'] = "=SUM(B5:E5)"
ws2['B7'].font = Font(bold=True)

ws2['A9'] = "⚠️ UBAH BOBOT DI BARIS 5 (Total harus = 1)"
ws2['A9'].font = Font(bold=True, color="FF0000")

# Sheet 3: Normalisasi
ws3 = wb.create_sheet("3. Normalisasi")
ws3['A1'] = "MATRIKS TERNORMALISASI (R)"
ws3['A1'].font = title_font
ws3['A2'] = "Rumus: rij = xij / √(Σxij²)"
ws3['A2'].font = Font(italic=True)

headers = ['Alternatif', 'CPU (C1)', 'RAM (C2)', 'Disk I/O (C3)', 'Harga (C4)']
for idx, h in enumerate(headers, 1):
    cell = ws3.cell(4, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Normalization formulas
for i in range(20):
    row = i + 5
    ws3.cell(row, 1, f"A{i+1}")
    
    # CPU normalization
    ws3.cell(row, 2, f"='1. Data Input'!D{i+4}/SQRT(SUMPRODUCT('1. Data Input'!$D$4:$D$23,'1. Data Input'!$D$4:$D$23))")
    
    # RAM normalization
    ws3.cell(row, 3, f"='1. Data Input'!E{i+4}/SQRT(SUMPRODUCT('1. Data Input'!$E$4:$E$23,'1. Data Input'!$E$4:$E$23))")
    
    # Disk I/O normalization
    ws3.cell(row, 4, f"='1. Data Input'!F{i+4}/SQRT(SUMPRODUCT('1. Data Input'!$F$4:$F$23,'1. Data Input'!$F$4:$F$23))")
    
    # Price normalization
    ws3.cell(row, 5, f"='1. Data Input'!G{i+4}/SQRT(SUMPRODUCT('1. Data Input'!$G$4:$G$23,'1. Data Input'!$G$4:$G$23))")

# Sheet 4: Normalisasi Terbobot
ws4 = wb.create_sheet("4. Normalisasi Terbobot")
ws4['A1'] = "MATRIKS TERNORMALISASI TERBOBOT (Y)"
ws4['A1'].font = title_font
ws4['A2'] = "Rumus: yij = wj × rij"
ws4['A2'].font = Font(italic=True)

for idx, h in enumerate(headers, 1):
    cell = ws4.cell(4, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Weighted normalization formulas
for i in range(20):
    row = i + 5
    ws4.cell(row, 1, f"A{i+1}")
    ws4.cell(row, 2, f"='3. Normalisasi'!B{row}*'2. Bobot & Kriteria'!$B$5")
    ws4.cell(row, 3, f"='3. Normalisasi'!C{row}*'2. Bobot & Kriteria'!$C$5")
    ws4.cell(row, 4, f"='3. Normalisasi'!D{row}*'2. Bobot & Kriteria'!$D$5")
    ws4.cell(row, 5, f"='3. Normalisasi'!E{row}*'2. Bobot & Kriteria'!$E$5")

# Sheet 5: Solusi Ideal
ws5 = wb.create_sheet("5. Solusi Ideal")
ws5['A1'] = "SOLUSI IDEAL POSITIF (A+) DAN NEGATIF (A-)"
ws5['A1'].font = title_font

ws5['A3'] = "Kriteria"
ws5['B3'] = "CPU (C1)"
ws5['C3'] = "RAM (C2)"
ws5['D3'] = "Disk I/O (C3)"
ws5['E3'] = "Harga (C4)"
for cell in ws5[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws5['A4'] = "A+ (Ideal Positif)"
ws5['B4'] = "=MAX('4. Normalisasi Terbobot'!B5:B24)"
ws5['C4'] = "=MAX('4. Normalisasi Terbobot'!C5:C24)"
ws5['D4'] = "=MAX('4. Normalisasi Terbobot'!D5:D24)"
ws5['E4'] = "=MIN('4. Normalisasi Terbobot'!E5:E24)"

ws5['A5'] = "A- (Ideal Negatif)"
ws5['B5'] = "=MIN('4. Normalisasi Terbobot'!B5:B24)"
ws5['C5'] = "=MIN('4. Normalisasi Terbobot'!C5:C24)"
ws5['D5'] = "=MIN('4. Normalisasi Terbobot'!D5:D24)"
ws5['E5'] = "=MAX('4. Normalisasi Terbobot'!E5:E24)"

ws5['A7'] = "Keterangan:"
ws5['A8'] = "• BENEFIT: A+ = MAX, A- = MIN"
ws5['A9'] = "• COST: A+ = MIN, A- = MAX"

# Sheet 6: Jarak & Score
ws6 = wb.create_sheet("6. Hasil TOPSIS")
ws6['A1'] = "HASIL PERHITUNGAN TOPSIS"
ws6['A1'].font = title_font

headers_final = ['Alternatif', 'Vendor', 'D+ (Jarak ke A+)', 'D- (Jarak ke A-)', 'Score', 'Rank']
for idx, h in enumerate(headers_final, 1):
    cell = ws6.cell(3, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Distance and score formulas
for i in range(20):
    row = i + 4
    ws6.cell(row, 1, f"A{i+1}")
    ws6.cell(row, 2, f"='1. Data Input'!B{row}")
    
    # D+ formula
    ws6.cell(row, 3, f"=SQRT(SUMPRODUCT(('4. Normalisasi Terbobot'!B{i+5}:E{i+5}-'5. Solusi Ideal'!$B$4:$E$4)^2))")
    
    # D- formula
    ws6.cell(row, 4, f"=SQRT(SUMPRODUCT(('4. Normalisasi Terbobot'!B{i+5}:E{i+5}-'5. Solusi Ideal'!$B$5:$E$5)^2))")
    
    # Score formula
    ws6.cell(row, 5, f"=D{row}/(C{row}+D{row})")
    
    # Rank formula
    ws6.cell(row, 6, f"=RANK(E{row},$E$4:$E$23,0)")

ws6['A26'] = "Rumus Score TOPSIS:"
ws6['A27'] = "Score = D- / (D+ + D-)"
ws6['A28'] = "Semakin tinggi score (mendekati 1), semakin baik alternatif"

# Adjust column widths
for ws in wb.worksheets:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

# Save
wb.save('TOPSIS_Rumus_Otomatis.xlsx')
print("✅ File Excel dengan rumus lengkap berhasil dibuat: TOPSIS_Rumus_Otomatis.xlsx")
print("\n📊 Cara Pakai:")
print("  1. Buka sheet '1. Data Input'")
print("  2. Ubah nilai di kolom KUNING (CPU, RAM, Disk I/O, Harga)")
print("  3. Ubah bobot di sheet '2. Bobot & Kriteria' jika perlu")
print("  4. Semua sheet lain akan OTOMATIS UPDATE!")
print("\n✨ Semua menggunakan FORMULA EXCEL, bukan nilai statis")
