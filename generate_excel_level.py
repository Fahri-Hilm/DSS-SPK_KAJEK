import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load data
df = pd.read_csv('No-Vendor-NamaPaketPlan-CPU-RAM-DiskIOSpeed-HargaBulanUSD.csv')

# Extract values
df['CPU_val'] = df['CPU'].apply(lambda x: int(re.search(r'(\d+)', x).group(1)))
df['RAM_val'] = df['RAM'].apply(lambda x: int(re.search(r'(\d+)', x).group(1)))

def extract_disk_io(val):
    speed_match = re.search(r'(\d+(?:\.\d+)?)\s*(Gbps|Mbps)', val)
    if speed_match:
        speed = float(speed_match.group(1))
        if 'Gbps' in speed_match.group(2):
            speed *= 1000
    else:
        speed = 100.0
    return speed / 8

df['DiskIO_val'] = df['Disk I/O Speed'].apply(extract_disk_io)
df['Price_val'] = df['Harga/Bulan (USD)'].apply(lambda x: float(re.search(r'[\$\s]*([\d.]+)', x).group(1)))

# Level functions
def get_cpu_level(val):
    if val <= 2: return 1
    elif val <= 4: return 2
    elif val <= 6: return 3
    elif val <= 8: return 4
    else: return 5

def get_ram_level(val):
    if val <= 2: return 1
    elif val <= 4: return 2
    elif val <= 8: return 3
    elif val <= 16: return 4
    else: return 5

def get_diskio_level(val):
    if val <= 200: return 1
    elif val <= 400: return 2
    elif val <= 600: return 3
    elif val <= 800: return 4
    else: return 5

def get_price_level(val):
    if val <= 20: return 1
    elif val <= 50: return 2
    elif val <= 100: return 3
    elif val <= 200: return 4
    else: return 5

df['CPU_Level'] = df['CPU_val'].apply(get_cpu_level)
df['RAM_Level'] = df['RAM_val'].apply(get_ram_level)
df['DiskIO_Level'] = df['DiskIO_val'].apply(get_diskio_level)
df['Price_Level'] = df['Price_val'].apply(get_price_level)

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E78")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Sheet 0: Panduan Level
ws0 = wb.create_sheet("0. Panduan Level")
ws0['A1'] = "PANDUAN LEVEL PARAMETER (1-5)"
ws0['A1'].font = Font(bold=True, size=16, color="1F4E78")

ws0['A3'] = "CPU (BENEFIT)"
ws0['A3'].font = Font(bold=True, size=12)
ws0['A4'] = "Level"
ws0['B4'] = "Range"
ws0['C4'] = "Kategori"
for cell in ws0['A4:C4'][0]:
    cell.fill = header_fill
    cell.font = header_font

data = [
    [1, "1-2 Core", "Sangat Rendah"],
    [2, "3-4 Core", "Rendah"],
    [3, "5-6 Core", "Sedang"],
    [4, "7-8 Core", "Tinggi"],
    [5, "9+ Core", "Sangat Tinggi"]
]
for i, row in enumerate(data, 5):
    ws0.cell(i, 1, row[0])
    ws0.cell(i, 2, row[1])
    ws0.cell(i, 3, row[2])

ws0['A11'] = "RAM (BENEFIT)"
ws0['A11'].font = Font(bold=True, size=12)
ws0['A12'] = "Level"
ws0['B12'] = "Range"
ws0['C12'] = "Kategori"
for cell in ws0['A12:C12'][0]:
    cell.fill = header_fill
    cell.font = header_font

data = [
    [1, "1-2 GB", "Sangat Rendah"],
    [2, "3-4 GB", "Rendah"],
    [3, "5-8 GB", "Sedang"],
    [4, "9-16 GB", "Tinggi"],
    [5, "17+ GB", "Sangat Tinggi"]
]
for i, row in enumerate(data, 13):
    ws0.cell(i, 1, row[0])
    ws0.cell(i, 2, row[1])
    ws0.cell(i, 3, row[2])

ws0['E3'] = "Disk I/O (BENEFIT)"
ws0['E3'].font = Font(bold=True, size=12)
ws0['E4'] = "Level"
ws0['F4'] = "Range"
ws0['G4'] = "Kategori"
for cell in ws0['E4:G4'][0]:
    cell.fill = header_fill
    cell.font = header_font

data = [
    [1, "100-200 MB/s", "Sangat Rendah"],
    [2, "201-400 MB/s", "Rendah"],
    [3, "401-600 MB/s", "Sedang"],
    [4, "601-800 MB/s", "Tinggi"],
    [5, "801+ MB/s", "Sangat Tinggi"]
]
for i, row in enumerate(data, 5):
    ws0.cell(i, 5, row[0])
    ws0.cell(i, 6, row[1])
    ws0.cell(i, 7, row[2])

ws0['E11'] = "Harga (COST)"
ws0['E11'].font = Font(bold=True, size=12)
ws0['E12'] = "Level"
ws0['F12'] = "Range"
ws0['G12'] = "Kategori"
for cell in ws0['E12:G12'][0]:
    cell.fill = header_fill
    cell.font = header_font

data = [
    [1, "$5-$20", "Sangat Murah"],
    [2, "$21-$50", "Murah"],
    [3, "$51-$100", "Sedang"],
    [4, "$101-$200", "Mahal"],
    [5, "$201+", "Sangat Mahal"]
]
for i, row in enumerate(data, 13):
    ws0.cell(i, 5, row[0])
    ws0.cell(i, 6, row[1])
    ws0.cell(i, 7, row[2])

# Sheet 1: Input Level
ws1 = wb.create_sheet("1. Input Level")
ws1['A1'] = "INPUT LEVEL (1-5) - UBAH DI KOLOM KUNING"
ws1['A1'].font = title_font

headers = ['No', 'Vendor', 'Nama Paket', 'CPU Lvl', 'RAM Lvl', 'I/O Lvl', 'Harga Lvl']
for idx, h in enumerate(headers, 1):
    cell = ws1.cell(3, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for idx, row in df.iterrows():
    ws1.cell(idx+4, 1, row['No'])
    ws1.cell(idx+4, 2, row['Vendor'])
    ws1.cell(idx+4, 3, row['Nama Paket (Plan)'])
    
    # Level input (yellow)
    for col in [4, 5, 6, 7]:
        cell = ws1.cell(idx+4, col)
        cell.fill = yellow_fill
    
    ws1.cell(idx+4, 4, row['CPU_Level'])
    ws1.cell(idx+4, 5, row['RAM_Level'])
    ws1.cell(idx+4, 6, row['DiskIO_Level'])
    ws1.cell(idx+4, 7, row['Price_Level'])

ws1['A25'] = "⚠️ INPUT LEVEL 1-5 DI KOLOM KUNING (Lihat sheet '0. Panduan Level')"
ws1['A25'].font = Font(bold=True, color="FF0000", size=12)

# Sheet 2: Konversi ke Nilai
ws2 = wb.create_sheet("2. Konversi ke Nilai")
ws2['A1'] = "KONVERSI LEVEL KE NILAI AKTUAL (OTOMATIS)"
ws2['A1'].font = title_font

headers = ['No', 'Vendor', 'CPU', 'RAM', 'Disk I/O', 'Harga']
for idx, h in enumerate(headers, 1):
    cell = ws2.cell(3, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for i in range(20):
    row = i + 4
    ws2.cell(row, 1, i+1)
    ws2.cell(row, 2, f"='1. Input Level'!B{row}")
    
    # CPU conversion: 1→2, 2→4, 3→6, 4→8, 5→10
    ws2.cell(row, 3, f"=CHOOSE('1. Input Level'!D{row},2,4,6,8,10)")
    
    # RAM conversion: 1→2, 2→4, 3→8, 4→16, 5→32
    ws2.cell(row, 4, f"=CHOOSE('1. Input Level'!E{row},2,4,8,16,32)")
    
    # Disk I/O conversion: 1→150, 2→300, 3→500, 4→700, 5→1000
    ws2.cell(row, 5, f"=CHOOSE('1. Input Level'!F{row},150,300,500,700,1000)")
    
    # Price conversion: 1→15, 2→35, 3→75, 4→150, 5→250
    ws2.cell(row, 6, f"=CHOOSE('1. Input Level'!G{row},15,35,75,150,250)")

# Sheet 3: Bobot
ws3 = wb.create_sheet("3. Bobot & Kriteria")
ws3['A1'] = "BOBOT DAN TIPE KRITERIA"
ws3['A1'].font = title_font

ws3['A3'] = "Kriteria"
ws3['B3'] = "CPU (C1)"
ws3['C3'] = "RAM (C2)"
ws3['D3'] = "Disk I/O (C3)"
ws3['E3'] = "Harga (C4)"
for cell in ws3[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws3['A4'] = "Tipe"
ws3['B4'] = "BENEFIT"
ws3['C4'] = "BENEFIT"
ws3['D4'] = "BENEFIT"
ws3['E4'] = "COST"

ws3['A5'] = "Bobot (W)"
for col in ['B5', 'C5', 'D5', 'E5']:
    ws3[col] = 0.25
    ws3[col].fill = yellow_fill

ws3['A7'] = "Total Bobot:"
ws3['B7'] = "=SUM(B5:E5)"
ws3['B7'].font = Font(bold=True)

# Sheet 4: Normalisasi
ws4 = wb.create_sheet("4. Normalisasi")
ws4['A1'] = "MATRIKS TERNORMALISASI (R)"
ws4['A1'].font = title_font
ws4['A2'] = "Rumus: rij = xij / √(Σxij²)"
ws4['A2'].font = Font(italic=True)

headers = ['Alternatif', 'CPU (C1)', 'RAM (C2)', 'Disk I/O (C3)', 'Harga (C4)']
for idx, h in enumerate(headers, 1):
    cell = ws4.cell(4, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for i in range(20):
    row = i + 5
    ws4.cell(row, 1, f"A{i+1}")
    ws4.cell(row, 2, f"='2. Konversi ke Nilai'!C{i+4}/SQRT(SUMPRODUCT('2. Konversi ke Nilai'!$C$4:$C$23,'2. Konversi ke Nilai'!$C$4:$C$23))")
    ws4.cell(row, 3, f"='2. Konversi ke Nilai'!D{i+4}/SQRT(SUMPRODUCT('2. Konversi ke Nilai'!$D$4:$D$23,'2. Konversi ke Nilai'!$D$4:$D$23))")
    ws4.cell(row, 4, f"='2. Konversi ke Nilai'!E{i+4}/SQRT(SUMPRODUCT('2. Konversi ke Nilai'!$E$4:$E$23,'2. Konversi ke Nilai'!$E$4:$E$23))")
    ws4.cell(row, 5, f"='2. Konversi ke Nilai'!F{i+4}/SQRT(SUMPRODUCT('2. Konversi ke Nilai'!$F$4:$F$23,'2. Konversi ke Nilai'!$F$4:$F$23))")

# Sheet 5: Normalisasi Terbobot
ws5 = wb.create_sheet("5. Normalisasi Terbobot")
ws5['A1'] = "MATRIKS TERNORMALISASI TERBOBOT (Y)"
ws5['A1'].font = title_font
ws5['A2'] = "Rumus: yij = wj × rij"
ws5['A2'].font = Font(italic=True)

for idx, h in enumerate(headers, 1):
    cell = ws5.cell(4, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for i in range(20):
    row = i + 5
    ws5.cell(row, 1, f"A{i+1}")
    ws5.cell(row, 2, f"='4. Normalisasi'!B{row}*'3. Bobot & Kriteria'!$B$5")
    ws5.cell(row, 3, f"='4. Normalisasi'!C{row}*'3. Bobot & Kriteria'!$C$5")
    ws5.cell(row, 4, f"='4. Normalisasi'!D{row}*'3. Bobot & Kriteria'!$D$5")
    ws5.cell(row, 5, f"='4. Normalisasi'!E{row}*'3. Bobot & Kriteria'!$E$5")

# Sheet 6: Solusi Ideal
ws6 = wb.create_sheet("6. Solusi Ideal")
ws6['A1'] = "SOLUSI IDEAL POSITIF (A+) DAN NEGATIF (A-)"
ws6['A1'].font = title_font

ws6['A3'] = "Kriteria"
ws6['B3'] = "CPU (C1)"
ws6['C3'] = "RAM (C2)"
ws6['D3'] = "Disk I/O (C3)"
ws6['E3'] = "Harga (C4)"
for cell in ws6[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws6['A4'] = "A+ (Ideal Positif)"
ws6['B4'] = "=MAX('5. Normalisasi Terbobot'!B5:B24)"
ws6['C4'] = "=MAX('5. Normalisasi Terbobot'!C5:C24)"
ws6['D4'] = "=MAX('5. Normalisasi Terbobot'!D5:D24)"
ws6['E4'] = "=MIN('5. Normalisasi Terbobot'!E5:E24)"

ws6['A5'] = "A- (Ideal Negatif)"
ws6['B5'] = "=MIN('5. Normalisasi Terbobot'!B5:B24)"
ws6['C5'] = "=MIN('5. Normalisasi Terbobot'!C5:C24)"
ws6['D5'] = "=MIN('5. Normalisasi Terbobot'!D5:D24)"
ws6['E5'] = "=MAX('5. Normalisasi Terbobot'!E5:E24)"

# Sheet 7: Hasil
ws7 = wb.create_sheet("7. Hasil TOPSIS")
ws7['A1'] = "HASIL PERHITUNGAN TOPSIS"
ws7['A1'].font = title_font

headers_final = ['Alternatif', 'Vendor', 'D+ (Jarak ke A+)', 'D- (Jarak ke A-)', 'Score', 'Rank']
for idx, h in enumerate(headers_final, 1):
    cell = ws7.cell(3, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for i in range(20):
    row = i + 4
    ws7.cell(row, 1, f"A{i+1}")
    ws7.cell(row, 2, f"='1. Input Level'!B{row}")
    ws7.cell(row, 3, f"=SQRT(SUMPRODUCT(('5. Normalisasi Terbobot'!B{i+5}:E{i+5}-'6. Solusi Ideal'!$B$4:$E$4)^2))")
    ws7.cell(row, 4, f"=SQRT(SUMPRODUCT(('5. Normalisasi Terbobot'!B{i+5}:E{i+5}-'6. Solusi Ideal'!$B$5:$E$5)^2))")
    ws7.cell(row, 5, f"=D{row}/(C{row}+D{row})")
    ws7.cell(row, 6, f"=RANK(E{row},$E$4:$E$23,0)")

ws7['A26'] = "Rumus Score TOPSIS: Score = D- / (D+ + D-)"
ws7['A27'] = "Semakin tinggi score (mendekati 1), semakin baik alternatif"

# Adjust column widths
for ws in wb.worksheets:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

wb.save('TOPSIS_Input_Level.xlsx')
print("✅ File Excel dengan input LEVEL berhasil dibuat: TOPSIS_Input_Level.xlsx")
print("\n📊 Cara Pakai:")
print("  1. Lihat sheet '0. Panduan Level' untuk referensi level 1-5")
print("  2. Buka sheet '1. Input Level'")
print("  3. Ubah LEVEL (1-5) di kolom KUNING")
print("  4. Sheet '2. Konversi ke Nilai' otomatis convert level ke nilai")
print("  5. Semua perhitungan TOPSIS otomatis update!")
print("\n✨ Input cukup angka 1-5, sistem otomatis konversi!")
