import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load data
df = pd.read_csv('No-Vendor-NamaPaketPlan-CPU-RAM-DiskIOSpeed-HargaBulanUSD.csv')

# Extract values
df['CPU_val'] = df['CPU'].apply(lambda x: int(re.search(r'(\d+)', x).group(1)))
df['RAM_val'] = df['RAM'].apply(lambda x: int(re.search(r'(\d+)', x).group(1)))

def extract_disk_io(val):
    speed_match = re.search(r'(\d+(?:\.\d+)?)\s*(Gbps|Mbps)', val)
    if speed_match:
        speed = float(speed_match.group(1))
        if 'Gbps' in speed_match.group(2):
            speed *= 1000
    else:
        speed = 100.0
    return speed / 8

df['DiskIO_val'] = df['Disk I/O Speed'].apply(extract_disk_io)
df['Price_val'] = df['Harga/Bulan (USD)'].apply(lambda x: float(re.search(r'[\$\s]*([\d.]+)', x).group(1)))

# Level functions
def get_cpu_level(val):
    if val <= 2: return 1
    elif val <= 4: return 2
    elif val <= 6: return 3
    elif val <= 8: return 4
    else: return 5

def get_ram_level(val):
    if val <= 2: return 1
    elif val <= 4: return 2
    elif val <= 8: return 3
    elif val <= 16: return 4
    else: return 5

def get_diskio_level(val):
    if val <= 200: return 1
    elif val <= 400: return 2
    elif val <= 600: return 3
    elif val <= 800: return 4
    else: return 5

def get_price_level(val):
    if val <= 20: return 1
    elif val <= 50: return 2
    elif val <= 100: return 3
    elif val <= 200: return 4
    else: return 5

df['CPU_Level'] = df['CPU_val'].apply(get_cpu_level)
df['RAM_Level'] = df['RAM_val'].apply(get_ram_level)
df['DiskIO_Level'] = df['DiskIO_val'].apply(get_diskio_level)
df['Price_Level'] = df['Price_val'].apply(get_price_level)

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E78")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Sheet 1: Data Asli
ws1 = wb.create_sheet("1. Data Asli")
ws1['A1'] = "DATA ASLI ALTERNATIF"
ws1['A1'].font = title_font

data_cols = ['No', 'Vendor', 'Nama Paket (Plan)', 'CPU_val', 'CPU_Level', 'RAM_val', 'RAM_Level', 
             'DiskIO_val', 'DiskIO_Level', 'Price_val', 'Price_Level']
for idx, col in enumerate(data_cols, 1):
    cell = ws1.cell(3, idx, col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for idx, row in df.iterrows():
    ws1.cell(idx+4, 1, row['No'])
    ws1.cell(idx+4, 2, row['Vendor'])
    ws1.cell(idx+4, 3, row['Nama Paket (Plan)'])
    ws1.cell(idx+4, 4, row['CPU_val'])
    ws1.cell(idx+4, 5, row['CPU_Level'])
    ws1.cell(idx+4, 6, row['RAM_val'])
    ws1.cell(idx+4, 7, row['RAM_Level'])
    ws1.cell(idx+4, 8, round(row['DiskIO_val'], 2))
    ws1.cell(idx+4, 9, row['DiskIO_Level'])
    ws1.cell(idx+4, 10, round(row['Price_val'], 2))
    ws1.cell(idx+4, 11, row['Price_Level'])

# Sheet 2: Matriks Keputusan
ws2 = wb.create_sheet("2. Matriks Keputusan")
ws2['A1'] = "MATRIKS KEPUTUSAN (X)"
ws2['A1'].font = title_font

headers = ['Alternatif', 'CPU (C1)', 'RAM (C2)', 'Disk I/O (C3)', 'Harga (C4)']
for idx, h in enumerate(headers, 1):
    cell = ws2.cell(3, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for idx, row in df.iterrows():
    ws2.cell(idx+4, 1, f"A{row['No']}")
    ws2.cell(idx+4, 2, row['CPU_val'])
    ws2.cell(idx+4, 3, row['RAM_val'])
    ws2.cell(idx+4, 4, round(row['DiskIO_val'], 2))
    ws2.cell(idx+4, 5, round(row['Price_val'], 2))

ws2['A26'] = "Tipe Kriteria:"
ws2['B26'] = "BENEFIT"
ws2['C26'] = "BENEFIT"
ws2['D26'] = "BENEFIT"
ws2['E26'] = "COST"
ws2['A27'] = "Bobot (W):"
ws2['B27'] = 0.25
ws2['C27'] = 0.25
ws2['D27'] = 0.25
ws2['E27'] = 0.25

# Sheet 3: Normalisasi
ws3 = wb.create_sheet("3. Normalisasi")
ws3['A1'] = "MATRIKS TERNORMALISASI (R)"
ws3['A1'].font = title_font
ws3['A2'] = "Rumus: rij = xij / √(Σxij²)"
ws3['A2'].font = Font(italic=True, size=10)

for idx, h in enumerate(headers, 1):
    cell = ws3.cell(4, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Calculate normalization
X = df[['CPU_val', 'RAM_val', 'DiskIO_val', 'Price_val']].values
sum_squares = np.sqrt((X**2).sum(axis=0))
X_norm = X / sum_squares
weights = np.array([0.25, 0.25, 0.25, 0.25])
X_weighted = X_norm * weights

for idx in range(len(df)):
    ws3.cell(idx+5, 1, f"A{df.iloc[idx]['No']}")
    for j in range(4):
        ws3.cell(idx+5, j+2, X_norm[idx, j])

# Sheet 4: Normalisasi Terbobot
ws4 = wb.create_sheet("4. Normalisasi Terbobot")
ws4['A1'] = "MATRIKS TERNORMALISASI TERBOBOT (Y)"
ws4['A1'].font = title_font
ws4['A2'] = "Rumus: yij = wj × rij"
ws4['A2'].font = Font(italic=True, size=10)

for idx, h in enumerate(headers, 1):
    cell = ws4.cell(4, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for idx in range(len(df)):
    ws4.cell(idx+5, 1, f"A{df.iloc[idx]['No']}")
    for j in range(4):
        ws4.cell(idx+5, j+2, X_weighted[idx, j])

# Sheet 5: Solusi Ideal
ws5 = wb.create_sheet("5. Solusi Ideal")
ws5['A1'] = "SOLUSI IDEAL POSITIF (A+) DAN NEGATIF (A-)"
ws5['A1'].font = title_font

ws5['A3'] = "Kriteria"
ws5['B3'] = "CPU (C1)"
ws5['C3'] = "RAM (C2)"
ws5['D3'] = "Disk I/O (C3)"
ws5['E3'] = "Harga (C4)"
for cell in ws5[3]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws5['A4'] = "A+ (Ideal Positif)"
ws5['B4'] = X_weighted[:, 0].max()
ws5['C4'] = X_weighted[:, 1].max()
ws5['D4'] = X_weighted[:, 2].max()
ws5['E4'] = X_weighted[:, 3].min()

ws5['A5'] = "A- (Ideal Negatif)"
ws5['B5'] = X_weighted[:, 0].min()
ws5['C5'] = X_weighted[:, 1].min()
ws5['D5'] = X_weighted[:, 2].min()
ws5['E5'] = X_weighted[:, 3].max()

ws5['A7'] = "Keterangan:"
ws5['A8'] = "• BENEFIT: A+ = MAX, A- = MIN"
ws5['A9'] = "• COST: A+ = MIN, A- = MAX"

# Sheet 6: Jarak & Score
ws6 = wb.create_sheet("6. Jarak & Score TOPSIS")
ws6['A1'] = "PERHITUNGAN JARAK DAN SCORE TOPSIS"
ws6['A1'].font = title_font

headers_final = ['Alternatif', 'Vendor', 'D+ (Jarak ke A+)', 'D- (Jarak ke A-)', 'Score', 'Rank']
for idx, h in enumerate(headers_final, 1):
    cell = ws6.cell(3, idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Calculate ideal solutions
ideal_pos = np.array([X_weighted[:, 0].max(), X_weighted[:, 1].max(), 
                      X_weighted[:, 2].max(), X_weighted[:, 3].min()])
ideal_neg = np.array([X_weighted[:, 0].min(), X_weighted[:, 1].min(), 
                      X_weighted[:, 2].min(), X_weighted[:, 3].max()])

D_pos = np.sqrt(((X_weighted - ideal_pos)**2).sum(axis=1))
D_neg = np.sqrt(((X_weighted - ideal_neg)**2).sum(axis=1))
scores = D_neg / (D_pos + D_neg)

for idx, row in df.iterrows():
    ws6.cell(idx+4, 1, f"A{row['No']}")
    ws6.cell(idx+4, 2, row['Vendor'])
    ws6.cell(idx+4, 3, D_pos[idx])
    ws6.cell(idx+4, 4, D_neg[idx])
    ws6.cell(idx+4, 5, scores[idx])
    ws6.cell(idx+4, 6, f"=RANK(E{idx+4},$E$4:$E$23,0)")

ws6['A26'] = "Rumus Score TOPSIS:"
ws6['A27'] = "Score = D- / (D+ + D-)"
ws6['A28'] = "Semakin tinggi score, semakin baik alternatif"

# Adjust column widths
for ws in wb.worksheets:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

# Save
wb.save('TOPSIS_Perhitungan_Lengkap.xlsx')
print("✅ File Excel berhasil dibuat: TOPSIS_Perhitungan_Lengkap.xlsx")
print("\n📊 Isi file:")
print("  1. Data Asli - Data mentah dengan level")
print("  2. Matriks Keputusan - Matriks X dengan bobot")
print("  3. Normalisasi - Matriks R (ternormalisasi)")
print("  4. Normalisasi Terbobot - Matriks Y")
print("  5. Solusi Ideal - A+ dan A-")
print("  6. Jarak & Score TOPSIS - Hasil akhir dengan ranking")
